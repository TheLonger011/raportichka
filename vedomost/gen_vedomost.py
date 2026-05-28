import sys, json, shutil
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter as gcl
from openpyxl.styles import Alignment, Font, Border, Side

STUDENT_START_ROW = 12
MAX_STUDENTS      = 24
TEMPLATE_SUMMARY  = 36
SUBJ_START_COL    = 3
TEMPLATE_N_SUBJ   = 6
SCORE_COLS        = 4
ABSENT_COLS       = 3

def _s(s): return Side(style=s) if s else Side()
def _b(left=None, right=None, top=None, bottom=None):
    return Border(left=_s(left), right=_s(right), top=_s(top), bottom=_s(bottom))

TNR10      = Font(name='Times New Roman', size=10)
TNR12      = Font(name='Times New Roman', size=12)
CENTER     = Alignment(horizontal='center', vertical='center', wrap_text=True)
CENTER_ROT = Alignment(horizontal='center', vertical='center', wrap_text=True, text_rotation=90)
LEFT_MID   = Alignment(horizontal='left', vertical='center', wrap_text=True)

def layout(n_subj):
    se  = SUBJ_START_COL + n_subj - 1
    sc  = SUBJ_START_COL + n_subj
    sce = sc + SCORE_COLS - 1
    ac  = sc + SCORE_COLS
    ace = ac + ABSENT_COLS - 1
    av  = ac + ABSENT_COLS
    return dict(sl=SUBJ_START_COL, se=se, sc=sc, sce=sce, ac=ac, ace=ace, av=av)

def unmerge_rows(ws, r1, r2):
    for mc in list(ws.merged_cells.ranges):
        if mc.min_row >= r1 and mc.max_row <= r2:
            ws.unmerge_cells(str(mc))

def adjust_columns(ws, n_subj):
    diff     = n_subj - TEMPLATE_N_SUBJ
    ins_at   = SUBJ_START_COL + TEMPLATE_N_SUBJ
    if diff > 0:
        ws.insert_cols(ins_at, diff)
        src_w = ws.column_dimensions[gcl(ins_at - 1)].width
        for c in range(ins_at, ins_at + diff):
            ws.column_dimensions[gcl(c)].width = src_w
    elif diff < 0:
        ws.delete_cols(ins_at + diff, -diff)

def rebuild_header(ws, n_subj, lyt):
    sl, se = lyt['sl'], lyt['se']
    sc, sce = lyt['sc'], lyt['sce']
    ac, ace = lyt['ac'], lyt['ace']
    av      = lyt['av']

    ws.merge_cells(start_row=7, start_column=1, end_row=11, end_column=2)
    c = ws.cell(7,1); c.value='ФИО студента'; c.font=TNR10; c.alignment=CENTER
    c.border=_b('medium','medium','medium','medium')

    ws.merge_cells(start_row=7, start_column=sl, end_row=10, end_column=se)
    c = ws.cell(7,sl); c.value='Дисциплины '; c.font=TNR10; c.alignment=CENTER
    c.border=_b('medium',None,'medium','medium')
    for col in range(sl+1, se+1):
        ws.cell(7,col).border=_b(top='medium')

    ws.merge_cells(start_row=7, start_column=sc, end_row=9, end_column=sce)
    c = ws.cell(7,sc); c.value='Количество оценок'; c.font=TNR10; c.alignment=CENTER
    c.border=_b('medium','medium','medium',None)

    ws.merge_cells(start_row=7, start_column=ac, end_row=9, end_column=ace)
    c = ws.cell(7,ac); c.value='Кол-во час.\n Пропущенных занятий'; c.font=TNR10; c.alignment=CENTER
    c.border=_b(None,'medium','medium',None)

    ws.merge_cells(start_row=7, start_column=av, end_row=9, end_column=av)
    c = ws.cell(7,av); c.value='Средний балл студента'; c.font=TNR10; c.alignment=CENTER
    c.border=_b('medium','medium','medium',None)

    score_labels  = ['отлично','хорошо','удовлетворительно','неудовлетворительно']
    absent_labels = ['всего','по уважительной\nпричине','без уважительной причины']

    for i in range(n_subj):
        c = ws.cell(11, sl+i); c.font=TNR10; c.alignment=CENTER_ROT
        c.border=_b('medium' if i==0 else 'thin', 'medium' if i==n_subj-1 else 'thin', None, 'thick')
    for i in range(SCORE_COLS):
        c = ws.cell(11, sc+i); c.value=score_labels[i]; c.font=TNR10; c.alignment=CENTER_ROT
        c.border=_b('medium' if i==0 else 'thin', 'medium' if i==SCORE_COLS-1 else 'thin', None, 'medium')
    for i in range(ABSENT_COLS):
        c = ws.cell(11, ac+i); c.value=absent_labels[i]; c.font=TNR10; c.alignment=CENTER_ROT
        c.border=_b(None if i==0 else 'thin', 'medium' if i==ABSENT_COLS-1 else 'thin',
                    'thin' if i==ABSENT_COLS-1 else None, 'medium')
    c = ws.cell(11,av); c.value=None; c.font=TNR10; c.alignment=CENTER_ROT
    c.border=_b('medium','medium','thin','medium')

def fill_students(ws, students, subjects, n_subj, lyt):
    sl, se = lyt['sl'], lyt['se']
    sc, ac, av = lyt['sc'], lyt['ac'], lyt['av']
    subj_id_to_idx = {s['id']: i for i, s in enumerate(subjects)}
    n_students = min(len(students), MAX_STUDENTS)

    for row in range(STUDENT_START_ROW, STUDENT_START_ROW + MAX_STUDENTS):
        for col in range(1, av + 1):
            ws.cell(row, col).value = None

    for i, st in enumerate(students[:n_students]):
        row = STUDENT_START_ROW + i
        sr  = f'{gcl(sl)}{row}:{gcl(se)}{row}'
        ws.cell(row,1).value = i + 1
        ws.cell(row,2).value = st['full_name']
        for sid_str, grade in st.get('grades', {}).items():
            idx = subj_id_to_idx.get(int(sid_str))
            if idx is not None and grade is not None:
                ws.cell(row, sl+idx).value = grade
        ws.cell(row, sc  ).value = f'=COUNTIF({sr},5)'
        ws.cell(row, sc+1).value = f'=COUNTIF({sr},4)'
        ws.cell(row, sc+2).value = f'=COUNTIF({sr},3)'
        ws.cell(row, sc+3).value = f'=COUNTIF({sr},2)'
        ws.cell(row, ac  ).value = st.get('absent_total', 0)
        ws.cell(row, ac+1).value = st.get('absent_excused', 0)
        ws.cell(row, ac+2).value = st.get('absent_unexcused', 0)
        ws.cell(row, av  ).value = f'=IFERROR(AVERAGE({sr}),"")'

    return n_students

def rebuild_summary(ws, n_subj, n_students, lyt, r36):
    sl, se = lyt['sl'], lyt['se']
    sc, sce = lyt['sc'], lyt['sce']
    ac, ace = lyt['ac'], lyt['ace']
    av      = lyt['av']

    st1 = STUDENT_START_ROW
    stN = STUDENT_START_ROW + n_students - 1

    r37, r38, r39, r40, r41 = r36+1, r36+2, r36+3, r36+4, r36+5
    r43, r44, r45, r46, r47 = r36+7, r36+8, r36+9, r36+10, r36+11
    r48, r49, r50           = r36+12, r36+13, r36+14

    for j in range(SCORE_COLS):
        col = sc+j
        ws.cell(r36, col).value = f'=SUM({gcl(col)}{st1}:{gcl(col)}{stN})'
    for j in range(ABSENT_COLS):
        col = ac+j
        ws.cell(r36, col).value = f'=SUM({gcl(col)}{st1}:{gcl(col)}{stN})'

    ws.merge_cells(start_row=r36, start_column=av, end_row=r41, end_column=av)

    ws.merge_cells(start_row=r37, start_column=1, end_row=r37, end_column=2)
    c = ws.cell(r37,1); c.value='Средний балл по дисциплине'; c.font=TNR10; c.alignment=LEFT_MID

    for j in range(n_subj):
        col = sl+j
        ws.cell(r37,col).value = f'=AVERAGE({gcl(col)}{st1}:{gcl(col)}{stN})'

    ws.merge_cells(start_row=r37, start_column=sc, end_row=r38, end_column=sce)
    c = ws.cell(r37,sc); c.value='Средний балл в группе'; c.font=TNR10; c.alignment=CENTER

    ws.merge_cells(start_row=r37, start_column=ac, end_row=r38, end_column=ace)
    c = ws.cell(r37,ac); c.value='Процент посещаемости'; c.font=TNR10; c.alignment=CENTER

    labels = ['Отлично','Хорошо','Удовлетворительно','Неудовлетворительно']
    grades = [5, 4, 3, 2]
    for i, (label, grade) in enumerate(zip(labels, grades)):
        row = r38 + i
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        c = ws.cell(row,1); c.value=label; c.font=TNR10; c.alignment=LEFT_MID
        for j in range(n_subj):
            col = sl+j
            ws.cell(row,col).value = f'=COUNTIF({gcl(col)}{st1}:{gcl(col)}{stN},{grade})'

    ws.merge_cells(start_row=r39, start_column=sc, end_row=r41, end_column=sce)
    c = ws.cell(r39,sc)
    c.value = f'=AVERAGE({gcl(sl)}{r37}:{gcl(se)}{r37})'
    c.font=TNR12; c.alignment=CENTER

    ws.merge_cells(start_row=r39, start_column=ac, end_row=r41, end_column=ace)
    c = ws.cell(r39,ac)
    c.value = f'=(100-({gcl(ac+2)}{r36}/(C{r46}*C{r47}))*100)'
    c.font=TNR12; c.alignment=CENTER

    sc0,sc1,sc2,sc3 = gcl(sc),gcl(sc+1),gcl(sc+2),gcl(sc+3)

    ws.cell(r43,2).value = 'Успеваемость по группе, %'
    ws.merge_cells(start_row=r43, start_column=3, end_row=r43, end_column=4)
    ws.cell(r43,3).value = f'=({sc0}{r36}*5+{sc1}{r36}*4+{sc2}{r36}*3)/({sc0}{r36}*5+{sc1}{r36}*4+{sc2}{r36}*3+{sc3}{r36}*2)*100'
    ws.cell(r43,3).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(r44,2).value = 'Качественная успеваемость, %'
    ws.merge_cells(start_row=r44, start_column=3, end_row=r44, end_column=4)
    ws.cell(r44,3).value = f'=({sc0}{r36}*5+{sc1}{r36}*4)/({sc0}{r36}*5+{sc1}{r36}*4+{sc2}{r36}*3+{sc3}{r36}*2)*100'
    ws.cell(r44,3).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(r45,2).value = 'Количество неуспевающих '
    ws.cell(r45,3).value = f'=COUNTIF({sc3}{st1}:{sc3}{stN},">0")'
    ws.cell(r46,2).value = 'Количество студентов в группе'
    ws.cell(r46,3).value = n_students
    ws.cell(r47,2).value = 'Количество учебных часов'
    ws.cell(r47,3).value = None

    return r46, r47, r48, r49

def main():
    data         = json.load(sys.stdin)
    subjects     = data['subjects']
    students     = data['students']
    n_subj       = len(subjects)
    total_hours  = data.get('total_hours', 120)
    head_teacher = data.get('head_teacher', '').strip()
    dept_head    = data.get('dept_head', '').strip()
    group_name   = data.get('group_name', '')
    month_label  = data.get('month_label', '')

    shutil.copy(data['template_path'], data['output_path'])
    wb = load_workbook(data['output_path'])
    ws = wb.active

    unmerge_rows(ws, 7, 11)
    unmerge_rows(ws, 36, 55)

    adjust_columns(ws, n_subj)
    lyt = layout(n_subj)

    rebuild_header(ws, n_subj, lyt)
    for i, subj in enumerate(subjects):
        ws.cell(11, lyt['sl']+i).value = subj['name']

    ws['B5'] = f'текущей аттестации за {month_label}'
    ws['B6'] = 'специальность' + ' '*145 + f'Курс   Группа {group_name}\n'

    n_students = fill_students(ws, students, subjects, n_subj, lyt)

    extra = MAX_STUDENTS - n_students
    if extra > 0:
        ws.delete_rows(STUDENT_START_ROW + n_students, extra)

    r36 = STUDENT_START_ROW + n_students
    r46, r47, r48, r49 = rebuild_summary(ws, n_subj, n_students, lyt, r36)

    ws.cell(r46, 3).value = n_students
    ws.cell(r47, 3).value = total_hours
    pad = ' ' * 70
    ws.cell(r48, 2).value = f'Классный руководитель{pad}{head_teacher}'
    ws.cell(r49, 2).value = f'Заведующий отделением{pad}{dept_head}'

    wb.save(data['output_path'])
    print(data['output_path'])

if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        import traceback; traceback.print_exc()
        print(f'ERROR: {e}', file=sys.stderr)
        sys.exit(1)